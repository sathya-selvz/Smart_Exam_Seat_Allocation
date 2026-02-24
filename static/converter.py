import json
from openpyxl import load_workbook


def excel_to_json(excel_file):
    # Load the Excel file using openpyxl
    workbook = load_workbook(excel_file)
    
    try:
        json_data = {}
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            # Convert sheet name to lowercase for consistent matching
            sheet_name_lower = sheet_name.lower().strip()
            
            # Get headers from first row
            headers = []
            for cell in sheet[1]:
                headers.append(cell.value)
            
            # Get data rows
            sheet_data = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_data = {}
                for idx, value in enumerate(row):
                    if idx < len(headers):
                        row_data[headers[idx]] = value
                sheet_data.append(row_data)
            
            # Use lowercase sheet name for consistent matching
            json_data[sheet_name_lower] = sheet_data
        
        return json_data
    finally:
        # Always close the workbook to release file locks
        workbook.close()
