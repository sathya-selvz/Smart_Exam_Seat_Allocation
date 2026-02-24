from flask import Flask, flash, render_template, request, redirect, url_for, jsonify, session, Markup
from datetime import datetime
from static.converter import excel_to_json
import os
import time
import math
import json
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from dotenv import load_dotenv
from pymongo import MongoClient
from flask_bcrypt import Bcrypt

# Load environment variables from .env file
load_dotenv()

# configuring flask

app = Flask(__name__)
app.debug = True
app.secret_key = os.getenv('SECRET_KEY', 'dev-secret-key-change-in-production')
bcrypt = Bcrypt(app)
app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(__file__), 'uploads')

# Get the MongoDB connection string from the environment variable
MONGO_URI = os.getenv("MONGO_URI")

# Connect to MongoDB
client = MongoClient(MONGO_URI)

# client = pymongo.MongoClient(
#     "mongodb://localhost:27017")

db = client.Studetails
usercollections = db.users
stucollections = db.student

# Create default admin user if it doesn't exist
def create_default_admin():
    try:
        if not usercollections.find_one({'username': 'admin'}):
            hashed_password = bcrypt.generate_password_hash('admin').decode('utf-8')
            usercollections.insert_one({'username': 'admin', 'password': hashed_password})
            print("✅ Default admin account created (username: admin, password: admin)")
        else:
            print("✅ Admin account already exists")
    except Exception as e:
        print(f"⚠️ Could not create default admin: {e}")

# Initialize default admin
create_default_admin()

# Helper function to generate formatted roll numbers
def generate_formatted_rollnum(year, sheet_name, serial_number):
    """
    Generate readable roll numbers based on year and department
    Examples:
    - 2nd year CE: 24CE001
    - 3rd year CSA: 23MCA013
    - 4th year EE: 22EE014
    """
    # Determine admission year prefix
    year_prefix = {
        "SecondYear": "24",    # Admitted in 2024
        "ThirdYear": "23",     # Admitted in 2023
        "FourthYear": "22"     # Admitted in 2022
    }
    
    # Department code mapping (uppercase)
    dept_mapping = {
        "csa": "MCA",
        "csb": "MCA",
        "ec": "EC",
        "ee": "MEE",
        "ce": "CE",
        "me": "ME",
        "mea": "ME",
        "meb": "ME",
        "ad": "AD",
        "mr": "MR",
        "rb": "RB"
    }
    
    year_code = year_prefix.get(year, "24")
    dept_code = dept_mapping.get(sheet_name.lower(), sheet_name.upper())
    
    # Format: YYDEPT### (e.g., 24CE001, 23MCA013) - all uppercase
    formatted_roll = f"{year_code}{dept_code}{serial_number:03d}"
    
    return formatted_roll

def detect_year_from_filename(filename):
    name = (filename or "").lower()
    if "second" in name or "2nd" in name or "year2" in name:
        return "SecondYear"
    if "third" in name or "3rd" in name or "year3" in name:
        return "ThirdYear"
    if "fourth" in name or "4th" in name or "year4" in name:
        return "FourthYear"
    return None

def normalize_student_excel(file_path, year_label):
    workbook = load_workbook(file_path)
    try:
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            roll_col = None
            for idx, cell in enumerate(sheet[1], start=1):
                value = str(cell.value).strip().lower() if cell.value is not None else ""
                if value in ("rollnum", "roll number", "rollno", "roll no"):
                    roll_col = idx
                    break
            if roll_col is None:
                continue

            for row_idx in range(2, sheet.max_row + 1):
                serial_number = row_idx - 1
                formatted_roll = generate_formatted_rollnum(year_label, sheet_name, serial_number)
                sheet.cell(row=row_idx, column=roll_col).value = formatted_roll
        workbook.save(file_path)
    finally:
        workbook.close()

# global variables
listy = []
filled = False
seating_data = {}  # Store seating results for immediate display
with open('static/dates.txt', 'r') as datefiles:
    dates = json.load(datefiles)

# routes
# homepage
@app.route('/')
def index():
    return render_template('home.html')

# signup page for admin
@app.route('/admin/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        # Check if the username already exists in the database
        if usercollections.find_one({'username': username}):
            flash('Username already exists', 'registration-error')
            return redirect(url_for('register'))
        
        else:
            # Hash the password before storing
            hashed_password = bcrypt.generate_password_hash(password).decode('utf-8')
            # If the username is unique, insert the new user into the database
            usercollections.insert_one(
                {'username': username, 'password': hashed_password})
            flash('Registration successful!', 'registration-success')
            return redirect(url_for('login'))
    else:
        return render_template('adminlogin.html')


# login page for admin
@app.route('/admin/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        # Retrieve the username and password from the form
        username = request.form['username']
        password = request.form['password']
        
        # Check if the username exists in the database
        user = usercollections.find_one({'username': username})
        
        if user and bcrypt.check_password_hash(user['password'], password):
            # If the user exists and password matches, store the username in the session
            session['username'] = username
            return redirect(url_for('admin'))
        else:
            flash('Invalid username or password', 'login-error')
            return redirect(url_for('login'))
    else:
        return render_template('adminlogin.html')


# main page of admin where he can choose the classes
@app.route('/admin')
def admin():
    return render_template('adminhome.html')


# When student enters their rollnumber
    # their corresponding seating is displayed
@app.route('/student', methods=['GET', 'POST'])
def student():
    if request.method == 'POST':
        roll = request.form['roll_num'].strip()
        
        # Try to find student by formatted rollnum (string) or original rollnum (int)
        student_data = stucollections.find_one({'rollnum': roll})
        if not student_data:
            # Try as integer if it's numeric
            try:
                student_data = stucollections.find_one({'rollnum': int(roll)})
            except (ValueError, TypeError):
                pass
        
        # Retrieve the seat number for the student
        seatnum = []
        if student_data is not None and 'seatnum' in student_data and student_data['seatnum'] is not None:
            seatnum = student_data['seatnum']
        return render_template('studentpage.html', roll_num=roll, seat_num=seatnum)
    else:
        return render_template('studentpage.html')


@app.route('/class', methods=['GET'])
def classchoose():
    return render_template('classavailable.html')


# page for uploading student details
@app.route('/uploaddata', methods=['GET'])
def uploadpage():
    return render_template('studentdataupload.html')

# when the data is submitted from /uploaddata or studentdataupload.html the data is processed here
# Here the data is checked and uploaded to the database
    # with sheetname as classname,year,classroom:which is the class they are going to be seated
# the data is also passed to "listy" for later usage in /seating
# finally the uploaded data is displayed in uploadeddata.html


@app.route('/upload', methods=['POST'])
def upload_file():
    files = request.files.getlist('files')

    if not files:
        flash('No files uploaded', 'error')
        return render_template('studentdataupload.html')

    if len(files) != 3:
        flash('Please select exactly 3 files (Second, Third, Fourth Year).', 'error')
        return render_template('studentdataupload.html')

    year_files = {"SecondYear": None, "ThirdYear": None, "FourthYear": None}
    unassigned = []
    for file in files:
        year_label = detect_year_from_filename(file.filename)
        if year_label and year_files[year_label] is None:
            year_files[year_label] = file
        else:
            unassigned.append(file)
    for year_label in ("SecondYear", "ThirdYear", "FourthYear"):
        if year_files[year_label] is None and unassigned:
            year_files[year_label] = unassigned.pop(0)
    if any(year_files[label] is None for label in year_files):
        flash('Could not determine all years from filenames. Please include Second/Third/Fourth in names.', 'error')
        return render_template('studentdataupload.html')

    file2 = year_files["SecondYear"]
    file3 = year_files["ThirdYear"]
    file4 = year_files["FourthYear"]

    if file2 and file2.filename:
        filename2 = secure_filename(file2.filename)
        file2.save(os.path.join(app.config['UPLOAD_FOLDER'], filename2))
        global data2
        normalize_student_excel(os.path.join(app.config['UPLOAD_FOLDER'], filename2), "SecondYear")
        data2 = excel_to_json(os.path.join(
            app.config['UPLOAD_FOLDER'], filename2))
    else:
        data2 = None

    if file3 and file3.filename:
        filename3 = secure_filename(file3.filename)
        file3.save(os.path.join(app.config['UPLOAD_FOLDER'], filename3))
        global data3
        normalize_student_excel(os.path.join(app.config['UPLOAD_FOLDER'], filename3), "ThirdYear")
        data3 = excel_to_json(os.path.join(
            app.config['UPLOAD_FOLDER'], filename3))
    else:
        data3 = None

    if file4 and file4.filename:
        filename4 = secure_filename(file4.filename)
        file4.save(os.path.join(app.config['UPLOAD_FOLDER'], filename4))
        global data4
        normalize_student_excel(os.path.join(app.config['UPLOAD_FOLDER'], filename4), "FourthYear")
        data4 = excel_to_json(os.path.join(
            app.config['UPLOAD_FOLDER'], filename4))
    else:
        data4 = None

    if data2 is not None:
        for sheet_name, sheet_data in data2.items():
            # Generate formatted roll numbers for each student
            formatted_data = []
            for idx, item in enumerate(sheet_data, start=1):
                formatted_roll = generate_formatted_rollnum("SecondYear", sheet_name, idx)
                formatted_data.append({
                    **item, 
                    "original_rollnum": item.get("rollnum"),  # Keep original if exists
                    "rollnum": formatted_roll,  # Use formatted roll number
                    "sheet_name": sheet_name, 
                    "Year": "SecondYear", 
                    "classroom": None
                })
            stucollections.insert_many(formatted_data)

    if data3 is not None:
        for sheet_name, sheet_data in data3.items():
            # Generate formatted roll numbers for each student
            formatted_data = []
            for idx, item in enumerate(sheet_data, start=1):
                formatted_roll = generate_formatted_rollnum("ThirdYear", sheet_name, idx)
                formatted_data.append({
                    **item, 
                    "original_rollnum": item.get("rollnum"),  # Keep original if exists
                    "rollnum": formatted_roll,  # Use formatted roll number
                    "sheet_name": sheet_name, 
                    "Year": "ThirdYear", 
                    "classroom": None
                })
            stucollections.insert_many(formatted_data)

    if data4 is not None:
        for sheet_name, sheet_data in data4.items():
            # Generate formatted roll numbers for each student
            formatted_data = []
            for idx, item in enumerate(sheet_data, start=1):
                formatted_roll = generate_formatted_rollnum("FourthYear", sheet_name, idx)
                formatted_data.append({
                    **item, 
                    "original_rollnum": item.get("rollnum"),  # Keep original if exists
                    "rollnum": formatted_roll,  # Use formatted roll number
                    "sheet_name": sheet_name, 
                    "Year": "FourthYear", 
                    "classroom": None
                })
            stucollections.insert_many(formatted_data)

    global listy
    listy = []
    details = []
    details = stucollections.aggregate(
        [{"$group": {"_id": "$subject", "ro": {"$push": "$rollnum"}}}])
    for i in details:
        listy.append(i)

    return render_template('uploadeddata.html', data2=data2, data3=data3, data4=data4)

# page for displaying the data via "GET" method
@app.route('/displaydata', methods=['GET'])
def display_data():
    return render_template('displaydata.html', data2=data2, data3=data3, data4=data4)

# here the timetable is uploaded via timetableupload.html
# the filename is checked
@app.route('/timetable', methods=['GET', 'POST'])
def timetable():
    if request.method == 'POST':
        # Retrieve uploaded files
        files = request.files.getlist('files')

        if not files:
            flash('No files uploaded', 'error')
            return render_template('timetableupload.html')

        if len(files) != 3:
            flash('Please select exactly 3 files (Second, Third, Fourth Year).', 'error')
            return render_template('timetableupload.html')

        year_files = {"SecondYear": None, "ThirdYear": None, "FourthYear": None}
        unassigned = []
        for file in files:
            year_label = detect_year_from_filename(file.filename)
            if year_label and year_files[year_label] is None:
                year_files[year_label] = file
            else:
                unassigned.append(file)
        for year_label in ("SecondYear", "ThirdYear", "FourthYear"):
            if year_files[year_label] is None and unassigned:
                year_files[year_label] = unassigned.pop(0)
        if any(year_files[label] is None for label in year_files):
            flash('Could not determine all years from filenames. Please include Second/Third/Fourth in names.', 'error')
            return render_template('timetableupload.html')

        file2 = year_files["SecondYear"]
        file3 = year_files["ThirdYear"]
        file4 = year_files["FourthYear"]

        # Check if file2 is uploaded
        if file2 and file2.filename:
            filename2 = secure_filename(file2.filename)
            file2.save(os.path.join(app.config['UPLOAD_FOLDER'], filename2))
            global timetable2
            timetable2 = excel_to_json(os.path.join(
                app.config['UPLOAD_FOLDER'], filename2))
        else:
            timetable2 = None

        # Check if file3 is uploaded
        if file3 and file3.filename:
            filename3 = secure_filename(file3.filename)
            file3.save(os.path.join(app.config['UPLOAD_FOLDER'], filename3))
            global timetable3
            timetable3 = excel_to_json(os.path.join(
                app.config['UPLOAD_FOLDER'], filename3))
        else:
            timetable3 = None

        # Check if file4 is uploaded
        if file4 and file4.filename:
            filename4 = secure_filename(file4.filename)
            file4.save(os.path.join(app.config['UPLOAD_FOLDER'], filename4))
            global timetable4
            timetable4 = excel_to_json(os.path.join(
                app.config['UPLOAD_FOLDER'], filename4))
        else:
            timetable4 = None

        # "Year" field is set to "SecondYear"
        # creates a list of the "_id" field values for those documents.
        # It then repeats this process for students in their third and fourth year of study,
        # Fetch student IDs for each year level

        second_year_students = stucollections.find({"Year": "SecondYear"})
        second_year_student_ids = [student["_id"]
                                   for student in second_year_students]
        third_year_students = stucollections.find({"Year": "ThirdYear"})
        third_year_student_ids = [student["_id"]
                                  for student in third_year_students]
        fourth_year_students = stucollections.find({"Year": "FourthYear"})
        fourth_year_student_ids = [student["_id"]
                                   for student in fourth_year_students]

        def get_subject_date(subject):
            date_value = subject.get("date")
            if date_value is not None:
                return date_value
            for key, value in subject.items():
                if isinstance(key, str) and key.strip().lower() == "date":
                    return value
            return None

        def get_subject_name(subject):
            subject_value = subject.get("subject")
            if subject_value is not None:
                return subject_value
            for key, value in subject.items():
                if isinstance(key, str) and key.strip().lower() == "subject":
                    return value
            return None

        # The code first checks if the timetable exists by checking if "timetable2" is not None.
        # If it does exist, the code iterates over the sheets in the timetable ("timetable2.items()"),
        # and for each subject in each sheet, it converts the "date" field to a string in the format '%d-%m-%Y'
        # using the "datetime.fromtimestamp()" and "strftime()" functions.
        # It then checks if the subject date is already in the "dates" list, and if not , adds it to the list.
        # The code then updates the "subject" field for each sheet in the "stucollections"
        # collection based on the sheet name, year level, and student IDs.
        # For each sheet, it uses the "update_many()" method to update the "subject" field of all documents in the collection
        # where the "sheet_name" field is equal to the current sheet, the "Year" field is equal to "SecondYear",
        # and the "_id" field is in the list of second-year student IDs retrieved earlier.

        # The updated "subject" field is set to the contents of the corresponding sheet in the "timetable2" dictionary,
        # which is accessed using the sheet name as the key(e.g., "timetable2["csa"]").
        # Update subjects in the "stucollections" collection based on the uploaded timetables

        if timetable2 is not None:
            for sheet_name, subjects in timetable2.items():
                for subject in subjects:
                    date_value = get_subject_date(subject)
                    if date_value is None:
                        continue
                    subject_name = get_subject_name(subject) or "Unknown"
                    # Handle date from openpyxl (returns datetime object)
                    if isinstance(date_value, datetime):
                        subject_date = date_value.strftime('%d-%m-%Y')
                    else:
                        # Fallback for other formats
                        subject_date = str(date_value)
                    subject['date'] = subject_date
                    subject['subject'] = subject_name
                    if subject_date not in dates:
                        dates.append(subject_date)
            
            # Department to timetable sheet mapping (handles section divisions)
            dept_timetable_mapping = {
                "csa": "cs", "csb": "cs",
                "ec": "ec", "ee": "ee", "ad": "ad", "ce": "ce",
                "me": "me", "mea": "me", "meb": "me",
                "mr": "mr", "rb": "rb",
                "mca": "mca"
            }

            # Update subjects for second year students using actual sheet names
            second_year_sheets = stucollections.distinct("sheet_name", {"Year": "SecondYear"})
            for dept_sheet in second_year_sheets:
                timetable_sheet = dept_timetable_mapping.get(dept_sheet, dept_sheet)
                if timetable_sheet in timetable2:
                    result = stucollections.update_many(
                        {"sheet_name": dept_sheet, "Year": "SecondYear",
                            "_id": {"$in": second_year_student_ids}},
                        {"$set": {"subject": timetable2[timetable_sheet]}}
                    )
                    print(f"Updated {result.modified_count} {dept_sheet} second year students with {timetable_sheet} subjects")

        if timetable3 is not None:
            for sheet_name, subjects in timetable3.items():
                for subject in subjects:
                    date_value = get_subject_date(subject)
                    if date_value is None:
                        continue
                    subject_name = get_subject_name(subject) or "Unknown"
                    # Handle date from openpyxl (returns datetime object)
                    if isinstance(date_value, datetime):
                        subject_date = date_value.strftime('%d-%m-%Y')
                    else:
                        # Fallback for other formats
                        subject_date = str(date_value)
                    subject['date'] = subject_date
                    subject['subject'] = subject_name
                    if subject_date not in dates:
                        dates.append(subject_date)
            
            # Department to timetable sheet mapping
            dept_timetable_mapping = {
                "csa": "cs", "csb": "cs",
                "ec": "ec", "ee": "ee", "ce": "ce",
                "mea": "me", "meb": "me", "me": "me",
                "mr": "mr", "ad": "ad", "rb": "rb",
                "mca": "mca"
            }

            # Update subjects for third year students using actual sheet names
            third_year_sheets = stucollections.distinct("sheet_name", {"Year": "ThirdYear"})
            for dept_sheet in third_year_sheets:
                timetable_sheet = dept_timetable_mapping.get(dept_sheet, dept_sheet)
                if timetable_sheet in timetable3:
                    result = stucollections.update_many(
                        {"sheet_name": dept_sheet, "Year": "ThirdYear",
                            "_id": {"$in": third_year_student_ids}},
                        {"$set": {"subject": timetable3[timetable_sheet]}}
                    )
                    print(f"Updated {result.modified_count} {dept_sheet} third year students with {timetable_sheet} subjects")

        if timetable4 is not None:
            for sheet_name, subjects in timetable4.items():
                for subject in subjects:
                    date_value = get_subject_date(subject)
                    if date_value is None:
                        continue
                    subject_name = get_subject_name(subject) or "Unknown"
                    # Handle date from openpyxl (returns datetime object)
                    if isinstance(date_value, datetime):
                        subject_date = date_value.strftime('%d-%m-%Y')
                    else:
                        # Fallback for other formats
                        subject_date = str(date_value)
                    subject['date'] = subject_date
                    subject['subject'] = subject_name
                    if subject_date not in dates:
                        dates.append(subject_date)
            
            # Department to timetable sheet mapping
            dept_timetable_mapping = {
                "csa": "cs", "csb": "cs",
                "ec": "ec", "ce": "ce", "ee": "ee",
                "me": "me", "mea": "me", "meb": "me",
                "mr": "mr",
                "mca": "mca"
            }

            # Update subjects for fourth year students using actual sheet names
            fourth_year_sheets = stucollections.distinct("sheet_name", {"Year": "FourthYear"})
            for dept_sheet in fourth_year_sheets:
                timetable_sheet = dept_timetable_mapping.get(dept_sheet, dept_sheet)
                if timetable_sheet in timetable4:
                    result = stucollections.update_many(
                        {"sheet_name": dept_sheet, "Year": "FourthYear",
                            "_id": {"$in": fourth_year_student_ids}},
                        {"$set": {"subject": timetable4[timetable_sheet]}}
                    )
                    print(f"Updated {result.modified_count} {dept_sheet} fourth year students with {timetable_sheet} subjects")

        with open('static/dates.txt', 'w') as f:
            json.dump(dates, f)
            flash('Upload successful', 'success')
        return render_template('timetableupload.html')
    else:
        flash('Upload failed', 'danger')
    return render_template('timetableupload.html')

# the timetable is fetched and displayed here
@app.route('/viewtimetable', methods=['GET'])
def view_timetable():
    # Fetch the documents from the MongoDB collection (only those with subject field)
    documents = stucollections.find(
        {'subject': {'$exists': True}}, {'sheet_name': 1, 'subject': 1, 'Year': 1})
    
    # Dictionary to store the timetable data
    timetables = {}
    for doc in documents:
        year = doc.get('Year')
        sheet_name = doc.get('sheet_name')
        subject = doc.get('subject')
        
        # Skip documents missing required fields
        if not year or not sheet_name or not subject:
            continue

        if year not in timetables:
            timetables[year] = {}

        if sheet_name not in timetables[year]:
            timetables[year][sheet_name] = []
        # Append the subject to the corresponding year and sheet_name in the timetable dictionary
        timetables[year][sheet_name].append(subject)
        
    # Convert the timetable dictionary to JSON format
    timetables = jsonify(timetables)
    return timetables


# unlike the /displaydata which displays the uploaded data
# this route fetches the uploaded data from the mongodb
@app.route('/viewdata', methods=['GET'])
def view_data():
    # Fetch the documents from the MongoDB collection
    documents = stucollections.find(
        {}, {'name': 1, 'rollnum': 1, 'sheet_name': 1, 'Year': 1})
    
    # List to store the retrieved data
    data = []
    
    for doc in documents:
        # Extract the relevant fields from each document, skip if missing required fields
        name = doc.get('name')
        rollnum = doc.get('rollnum')
        sheet_name = doc.get('sheet_name')
        year = doc.get('Year')
        
        # Only add documents that have all required fields
        if name and rollnum and sheet_name and year:
            data.append({
                'name': name,
                'rollnum': rollnum,
                'sheet_name': sheet_name,
                'Year': year
            })
        
    # Convert the data list to JSON format
    data = jsonify(data)
    return data


# here we are assigning the classname and seat num for each class
@app.route('/details', methods=['POST'])
def details():
    if request.method == 'POST':
        # Get the list of selected items from the form
        items = request.form.getlist('item[]')

        # List to store the details of selected classes
        class_data = []

        # Dictionary mapping class items to their details
        # Standardized: 4 columns × 6 rows = 24 seats per classroom
        class_details = {
            'ADM 303': {'class_name': 'ADM 303', 'column': 4, 'rows': 6},
            'ADM 304': {'class_name': 'ADM 304', 'column': 4, 'rows': 6},
            'ADM 305': {'class_name': 'ADM 305', 'column': 4, 'rows': 6},
            'ADM 306': {'class_name': 'ADM 306', 'column': 4, 'rows': 6},
            'ADM 307': {'class_name': 'ADM 307', 'column': 4, 'rows': 6},
            'ADM 308': {'class_name': 'ADM 308', 'column': 4, 'rows': 6},
            'ADM 309': {'class_name': 'ADM 309', 'column': 4, 'rows': 6},
            'ADM 310': {'class_name': 'ADM 310', 'column': 4, 'rows': 6},
            'ADM 311': {'class_name': 'ADM 311', 'column': 4, 'rows': 6},
            'EAB 206': {'class_name': 'EAB 206', 'column': 4, 'rows': 6},
            'EAB 306': {'class_name': 'EAB 306', 'column': 4, 'rows': 6},
            'EAB 401': {'class_name': 'EAB 401', 'column': 4, 'rows': 6},
            'EAB 304': {'class_name': 'EAB 304', 'column': 4, 'rows': 6},
            'EAB 303': {'class_name': 'EAB 303', 'column': 4, 'rows': 6},
            'EAB 104': {'class_name': 'EAB 104', 'column': 4, 'rows': 6},
            'EAB 103': {'class_name': 'EAB 103', 'column': 4, 'rows': 6},
            'EAB 203': {'class_name': 'EAB 203', 'column': 4, 'rows': 6},
            'EAB 204': {'class_name': 'EAB 204', 'column': 4, 'rows': 6},
            'WAB 206': {'class_name': 'WAB 206', 'column': 4, 'rows': 6},
            'WAB 105': {'class_name': 'WAB 105', 'column': 4, 'rows': 6},
            'WAB 107': {'class_name': 'WAB 107', 'column': 4, 'rows': 6},
            'WAB 207': {'class_name': 'WAB 207', 'column': 4, 'rows': 6},
            'WAB 212': {'class_name': 'WAB 212', 'column': 4, 'rows': 6},
            'WAB 210': {'class_name': 'WAB 210', 'column': 4, 'rows': 6},
            'WAB 211': {'class_name': 'WAB 211', 'column': 4, 'rows': 6},
            'WAB 205': {'class_name': 'WAB 205', 'column': 4, 'rows': 6},
            'WAB 305': {'class_name': 'WAB 305', 'column': 4, 'rows': 6},
            'WAB 303': {'class_name': 'WAB 303', 'column': 4, 'rows': 6},
            'WAB 403': {'class_name': 'WAB 403', 'column': 4, 'rows': 6},
            'WAB 405': {'class_name': 'WAB 405', 'column': 4, 'rows': 6},
            'EAB 415': {'class_name': 'EAB 415', 'column': 4, 'rows': 6},
            'EAB 416': {'class_name': 'EAB 416', 'column': 4, 'rows': 6},
            'WAB 412': {'class_name': 'WAB 412', 'column': 4, 'rows': 6},
            'EAB 310': {'class_name': 'EAB 310', 'column': 4, 'rows': 6},
        }

        for item in items:
            if item in class_details:
                class_data.append(class_details[item])
        # Write the class_data list to 'static/stuarrange.txt' file as JSON (compact format)
        with open('static/stuarrange.txt', 'w') as f:
            json.dump(class_data, f)

        global filled
        filled = False
        return render_template('classdetails.html', class_data=class_data)


# here the seating is done
# only two students can sit one bench but with different subjects as exam
# -issue-:this issue may arise when there is limited class and students with same subject maybe seated nearby
# using the skeleton file stuarrange.txt the students are seated into the classroom
# the timetable/date is noted . stuarrange.txt files which is the seating arrangement is generated for each day in the timetable

@app.route('/seating', methods=['GET'])
def seating():
    global filled
    
    # Check if 'stuarrange.txt' file doesn't exist, flash an error message and redirect to 'admin' route
    if not os.path.exists('static/stuarrange.txt'):
        flash('Choose Class', 'error')
        return redirect(url_for('admin'))
    
    if filled:
        with open('static/stuarrange.txt', 'r') as stufiles:
            stulist = json.load(stufiles)  # Load the JSON data from the file
        flash('Already generated', 'error')
        return redirect(url_for('admin'))
    
    else:
        # Check if students have subjects assigned (timetable uploaded)
        student_with_subject = stucollections.find_one({"subject": {"$exists": True, "$ne": None}})
        if not student_with_subject:
            flash('Please upload timetable first before generating seating!', 'error')
            return redirect(url_for('admin'))

        # Always reload dates to reflect latest timetable upload
        with open('static/dates.txt', 'r') as datefiles:
            dates = json.load(datefiles)
        
        #reset data to avoid redundancy
        stucollections.update_many({}, {"$unset": {"seatnum": ""}})
        
        for date in dates:
            # Build subject groups for this date directly from students
            subject_groups = {}
            students_with_exams = stucollections.find(
                {"subject": {"$elemMatch": {"date": date}}},
                {"rollnum": 1, "subject": 1}
            )
            for student in students_with_exams:
                rollnum = student.get("rollnum")
                for subj in student.get("subject", []):
                    if subj.get("date") == date:
                        key = subj.get("subject", "Unknown")
                        if key not in subject_groups:
                            subject_groups[key] = {"_id": subj, "ro": []}
                        subject_groups[key]["ro"].append(rollnum)

            listy = list(subject_groups.values())

            with open('static/stuarrange.txt', 'r') as stufiles:
                stulist = json.load(stufiles)
            
            # Extract year and department from roll numbers
            def normalize_dept(dept):
                """Convert 2-char departments to 3-char"""
                dept_mapping = {
                    "EE": "MEE",    # Electrical/Mechanical Engineering
                    "EC": "ECE",    # Electronics & Communication Engineering
                    "CE": "CIV",    # Civil Engineering
                    "ME": "MEC",    # Mechanical Engineering
                    "RB": "RBE",    # Robotics & Automation
                    "AD": "ADE",    # Additional
                    "MR": "MRS",    # Miscellaneous
                }
                return dept_mapping.get(dept, dept)  # Return mapped or original if not found
            
            def extract_year_dept(rollnum):
                """Extract year and department from roll number
                Format: YYDEPT### where dept can be 2-3 letters (EE, ME, MEE, RB, MCA, etc)
                """
                if len(rollnum) < 4:
                    return None, None
                
                year = rollnum[:2]  # First 2 digits (22, 23, 24)
                
                # Extract department (2-3 letters, stop at first digit)
                dept = ""
                for char in rollnum[2:]:
                    if char.isalpha():
                        dept += char
                    else:
                        break
                
                # Normalize 2-char departments to 3-char
                dept = normalize_dept(dept)
                
                return year, dept if dept else None
            
            # Build year -> department -> exam -> unique rollnums for this date
            year_dept_exam_map = {}
            seen_rollnums = set()
            for subject_item in listy:
                if subject_item["ro"] and len(subject_item["ro"]) > 0:
                    exam_name = subject_item.get("_id", {}).get("subject", "Unknown")
                    for rollnum in subject_item.get("ro", []):
                        if rollnum in seen_rollnums:
                            continue
                        seen_rollnums.add(rollnum)
                        year, dept = extract_year_dept(rollnum)
                        if not year or not dept:
                            continue
                        if year not in year_dept_exam_map:
                            year_dept_exam_map[year] = {}
                        if dept not in year_dept_exam_map[year]:
                            year_dept_exam_map[year][dept] = {}
                        if exam_name not in year_dept_exam_map[year][dept]:
                            year_dept_exam_map[year][dept][exam_name] = []
                        year_dept_exam_map[year][dept][exam_name].append(rollnum)

            # Sort roll numbers for stable ordering
            for year in year_dept_exam_map:
                for dept, exam_map in year_dept_exam_map[year].items():
                    for exam_name, rollnums in exam_map.items():
                        rollnums.sort()

            # Order years in descending order (24, 23, 22, etc.) for consistent placement
            year_order = sorted(year_dept_exam_map.keys(), reverse=True)
            
            # Debug: Print department distribution
            print("\n=== Department Distribution ===")
            for year in year_order:
                for dept in sorted(year_dept_exam_map[year].keys()):
                    total_students = sum(len(rollnums) for rollnums in year_dept_exam_map[year][dept].values())
                    print(f"Year {year}, Dept {dept}: {total_students} students")
            print("================================\n")

            # Build queues per year and department (optimized)
            from collections import deque
            year_dept_queues = {}
            for year in year_order:
                year_dept_queues[year] = {}
                dept_order = sorted(year_dept_exam_map[year].keys())
                for dept in dept_order:
                    exam_map = year_dept_exam_map[year][dept]
                    exam_names = sorted(exam_map.keys())
                    
                    # Convert to deques for O(1) popleft
                    exam_deques = {exam: deque(rollnums) for exam, rollnums in exam_map.items()}
                    
                    # Round-robin through exams efficiently
                    queue = []
                    while exam_deques:
                        for exam in list(exam_deques.keys()):
                            if exam_deques[exam]:
                                queue.append((exam_deques[exam].popleft(), exam))
                                if not exam_deques[exam]:
                                    del exam_deques[exam]
                    
                    year_dept_queues[year][dept] = queue
            
            # Batch seat assignments for performance
            seat_assignments = []

            # Build student list with round-robin across departments
            all_students = []
            
            for year in year_order:
                dept_order = sorted(year_dept_queues[year].keys())
                
                # Convert to deques for efficient popping
                dept_deques = {dept: deque(year_dept_queues[year][dept]) for dept in dept_order}
                
                # Round-robin through departments to distribute students evenly
                while dept_deques:
                    for dept in dept_order:  # Keep original order for consistency
                        if dept in dept_deques and dept_deques[dept]:
                            rollnum, exam_name = dept_deques[dept].popleft()
                            all_students.append((rollnum, exam_name, dept, year))
                            if not dept_deques[dept]:
                                del dept_deques[dept]
            
            student_idx = 0  # Track position in the ordered student list

            for i in stulist:
                i["a"] = []
                i["b"] = []
                i["c"] = []
                i["d"] = []
                i["e"] = []
                i["f"] = []
                i["g"] = []
                i["h"] = []
                class_name = i.get("class_name")

                rows = int(i["rows"])
                column_order = ["a", "b", "c", "d", "e", "f", "g", "h"]  # 8 columns

                # Fill ROW BY ROW (not column by column) to ensure different depts in each row
                for row_num in range(1, rows + 1):
                    for col_key in column_order:
                        if student_idx >= len(all_students):
                            break
                        
                        rollnum, exam_name, dept, year = all_students[student_idx]
                        
                        student_data = {
                            "rollnum": rollnum,
                            "dept": dept,
                            "exam": exam_name
                        }
                        i[col_key].append(student_data)

                        seatinfo = {
                            "date": date,
                            "seatnum": f"{col_key}{row_num}",
                            "classroom": class_name,
                            "subject": exam_name
                        }
                        seat_assignments.append({"rollnum": rollnum, "seatinfo": seatinfo})
                        
                        student_idx += 1
            
            # Batch database updates using bulk_write for performance
            from pymongo import UpdateOne
            student_seat_map = {}
            for assignment in seat_assignments:
                rollnum = assignment["rollnum"]
                if rollnum not in student_seat_map:
                    student_seat_map[rollnum] = []
                student_seat_map[rollnum].append(assignment["seatinfo"])
            
            # Use bulk_write for efficient database updates
            if student_seat_map:
                bulk_operations = [
                    UpdateOne(
                        {"rollnum": rollnum},
                        {"$addToSet": {"seatnum": {"$each": seats}}}
                    )
                    for rollnum, seats in student_seat_map.items()
                ]
                stucollections.bulk_write(bulk_operations, ordered=False)
            
            # Store seating data for immediate display
            global seating_data
            seating_data = {date: stulist}
            
            newlist = list(stulist)
            
            # Check if all students were seated
            # Count students from all_students queue that were not placed
            unseated_total = len(all_students) - student_idx
            
            if unseated_total > 0:
                # There are more students than available seats
                flash(f'Warning: {unseated_total} students could not be seated. Available capacity is insufficient.', 'danger')
                return render_template('classavailable.html', stunum=unseated_total)
            
            # Seating file output disabled (per request)
            filled = True  # Set 'filled' to True to indicate that seating is generated


        flash('Generated', 'success')
        # Display seating immediately after generation with in-memory data
        if seating_data:
            first_date = dates[0]
            class_list = seating_data.get(first_date, [])
            return render_template('viewseating.html', dates=dates, date_exams={}, initial_data=class_list)
        return render_template("adminhome.html")


@app.route('/viewseating', methods=['GET'])
def viewseating():
    global filled
    if not filled:
        flash('Firstly generate seating', 'error')
        return render_template("adminhome.html")
    
    # Load dates
    with open('static/dates.txt', 'r') as file:
        dates = json.load(file)
    
    # Get exam subjects for each date
    date_exams = {}
    for date in dates:
        # Query students who have exams on this date
        students_with_exams = stucollections.find(
            {"subject": {"$elemMatch": {"date": date}}},
            {"subject": 1}
        )
        
        # Collect unique subjects for this date
        subjects_set = set()
        for student in students_with_exams:
            if student.get('subject'):
                for subj in student['subject']:
                    if subj.get('date') == date:
                        subjects_set.add(subj.get('subject', 'Unknown'))
        
        date_exams[date] = list(subjects_set)
    
    return render_template('viewseating.html', dates=dates, date_exams=date_exams)
# Render the 'viewseating.html' template, passing the content of 'dates.txt' as the 'dates' variable
# Markup is used to mark the content as safe to render HTML tags, assuming the content contains HTML


@app.route('/viewseating/<path:name>', methods=['GET'])
def viewseating1(name):
    global filled, seating_data
    if filled:
        # First try to use in-memory seating data from recent generation
        if name in seating_data:
            class_list = seating_data[name]
            return json.dumps(class_list)
        
        # Fallback to database if data not in memory
        def normalize_dept(dept):
            dept_mapping = {
                "EE": "MEE",
                "EC": "ECE",
                "CE": "CIV",
                "ME": "MEC",
                "RB": "RBE",
                "AD": "ADE",
                "MR": "MRS",
            }
            return dept_mapping.get(dept, dept)

        def extract_dept(rollnum):
            if not rollnum or len(rollnum) < 4:
                return ""
            dept = ""
            for char in rollnum[2:]:
                if char.isalpha():
                    dept += char
                else:
                    break
            return normalize_dept(dept) if dept else ""

        class_map = {}
        students_with_seats = stucollections.find(
            {"seatnum": {"$elemMatch": {"date": name}}},
            {"rollnum": 1, "seatnum": 1}
        )

        for student in students_with_seats:
            rollnum = student.get("rollnum")
            dept = extract_dept(rollnum)
            for seatinfo in student.get("seatnum", []):
                if seatinfo.get("date") != name:
                    continue
                classroom = seatinfo.get("classroom")
                seatnum = str(seatinfo.get("seatnum", "")).lower()
                subject = seatinfo.get("subject", "Unknown")
                if not classroom or not seatnum or len(seatnum) < 2:
                    continue

                col_key = seatnum[0]
                if col_key not in ("a", "b", "c", "d", "e", "f", "g", "h"):
                    continue
                try:
                    seat_index = int(seatnum[1:]) - 1
                except ValueError:
                    continue

                if classroom not in class_map:
                    class_map[classroom] = {
                        "class_name": classroom,
                        "column": 8,
                        "rows": 6,
                        "a": [],
                        "b": [],
                        "c": [],
                        "d": [],
                        "e": [],
                        "f": [],
                        "g": [],
                        "h": []
                    }

                col_list = class_map[classroom][col_key]
                while len(col_list) <= seat_index:
                    col_list.append(None)

                col_list[seat_index] = {
                    "rollnum": rollnum,
                    "dept": dept,
                    "exam": subject
                }

        class_list = sorted(class_map.values(), key=lambda c: c["class_name"])
        return json.dumps(class_list)
    else:
        flash('Firstly generate seating', 'error')
        return render_template("adminhome.html")
    


# Resetting everything out
@app.route('/reset', methods=['GET'])
def reset():
    return render_template('reset.html')


@app.route('/reset/collections', methods=['GET'])
def reset_collections():
    global filled
    filled = False
    stucollections.drop()  # Drop the 'student' collection
    message = "Student data has been deleted."
    return render_template('reset.html', message=message)


@app.route('/reset/users', methods=['GET'])
def reset_users():
    usercollections.drop()  # Drop the 'users' collection
    message = "Users has been deleted."
    return render_template('reset.html', message=message)


@app.route('/reset/static', methods=['GET'])
def reset_static():
    folder_path = 'static'
    files = os.listdir(folder_path)  # Get a list of all files in the folder
    for file in files:
        if file.startswith("stuarrange"):
            # Get the full path of the file
            file_path = os.path.join(folder_path, file)
            os.remove(file_path)  # Remove the file from the folder
    message = "Static files have been reset."
    global filled
    filled = False
    return render_template('reset.html', message=message)


@app.route('/reset/uploads', methods=['GET'])
def reset_uploads():
    folder_path = 'uploads'
    try:
        files = os.listdir(folder_path)
        for file in files:
            file_path = os.path.join(folder_path, file)
            try:
                # Try to remove the file
                os.remove(file_path)
            except PermissionError:
                # If file is locked, wait and retry
                time.sleep(0.1)
                try:
                    os.remove(file_path)
                except Exception as e:
                    flash(f'Could not delete {file}: {str(e)}', 'warning')
                    continue
        message = "Uploads have been reset."
        return render_template('reset.html', message=message)
    except Exception as e:
        flash(f'Error during reset: {str(e)}', 'error')
        return render_template('reset.html', message=f'Error: {str(e)}')


@app.route('/reset/dates', methods=['GET'])
def reset_dates():
    folder_path = 'static'
    file_path = os.path.join(folder_path, 'dates.txt')
    with open(file_path, 'w') as file:
        file.write('[]')
    message = "Dates have been reset."
    return render_template('reset.html', message=message)

# main function
if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0')


